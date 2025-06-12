import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from django.conf import settings

def genereate_excel(data,purchase_number):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    # 输入一些数据
    sheet['A1']='供应商:'
    sheet['B1']=data[0]['supplier_name']
    sheet['A2']='联系人:'
    sheet['B2']=data[0]['supplier_contact_name']
    sheet['A3']='联系方式:'
    sheet['B3']=data[0]['supplier_contact_phone']
    sheet['A4']='采购员:'
    sheet['B4']=data[0]['username']

    sheet['F1']='采购单号:'
    sheet['G1']=data[0]['purchase_number']
    sheet['F2']='采购日期:'
    sheet['G2']=data[0]['purchase_date']
    sheet['F3']='币   种:'
    sheet['G3']=data[0]['currency']


    sheet['A5'] = ''

    columns = [
        '序号','品名', '规格', '单位', '数量', '单价',
        '总价', '备注'
    ]
    sheet.append(columns)
    index = 1
    total = 0
    for item in data:
        sheet.append([
            index,
            item["product_name"],
            item["product_size"],
            item["unit"],
            item["quantity"],
            item["unit_price"],
            item["total_price"],
            item["remark"],
        ])
        total = total + item["quantity"]
        index = index+1

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['H'].width = 20


    sheet[f'D{index+6}']='合计'
    sheet[f'E{index + 6}'] = total

    sheet.merge_cells('G2:H2')
    sheet.merge_cells('G3:H3')
    sheet['G2'].alignment = Alignment(
        horizontal='left',
        vertical='bottom'
    )
    sheet['G3'].alignment = Alignment(
        horizontal='left',
        vertical='bottom'
    )

    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0,
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row_num in range(0, index+7):
        sheet.row_dimensions[row_num].height = 20  # 设置行高为 30

    for row in sheet.iter_rows(min_row=6, max_row=index+5, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border

    sheet[f'D{index + 6}'].alignment = alignment
    sheet[f'E{index + 6}'].alignment = alignment

    excel_file_name = f'purchase_{purchase_number}.xlsx'
    file_path = os.path.join(settings.PDF_DIR, excel_file_name)
    workbook.save(file_path)
    # 确保文件存在
    # if not os.path.exists(file_path):
    #     return HttpResponse(status=404)
    # # 返回文件响应
    # responseFile = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f'shipping_{shippingid}.xlsx')
    # responseFile['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return excel_file_name