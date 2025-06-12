import os
import math
from urllib.parse import quote
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, FileResponse
from app.api_code import *
from app.models import ShippingTable, ShippingDetailTable, OrderDetailTable
from app.models import ProjectTable
from app.models import DriverTable
from app.models import UserTable
from app.models import CustomerTable
from loguru import logger
import pdfkit
from jinja2 import Template
from app.util.db_util import DBUtil
from app.views.user_table_view import UserTableView
from django.db.models import OuterRef, Subquery, Value, CharField
from django.db.models import F, Q, Value
from django.db.models import Aggregate
import io, zipfile, datetime, uuid
from django.conf import settings
import datetime, logging, urllib.parse
from openpyxl.styles import Alignment, Border, Side, Font
import urllib.parse, datetime
from typing import Union
from io import BytesIO
import urllib.parse, datetime
from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from typing import List, Dict, Tuple          # ← 加这一行
from io import BytesIO
from urllib.parse import quote
from django.http import StreamingHttpResponse, JsonResponse
import pdfkit, math, os
from jinja2 import Template
from django.conf import settings

logger = logging.getLogger(__name__)

RAW_SQL = """
SELECT
    sd.detailid_id                                                AS id,
    sd.shippingid_id                                              AS shippingid_id,
    st.delivery_date,
    st.shipping_address,
    GROUP_CONCAT(DISTINCT d.driver_name)                          AS driver_name,
    GROUP_CONCAT(DISTINCT d.licence_plate)                        AS licence_plate,
    GROUP_CONCAT(DISTINCT p.project_contact_name)                 AS project_contact_name,
    GROUP_CONCAT(DISTINCT p.project_contact_phone)                AS project_contact_phone,
    ot.project_name,
    od.customer_name_id,
    od.commodity_details,
    od.commodity_size,
    od.commodity_units,
    od.control_no,
    od.unit_weight,
    sd.shipping_quantity,
    sd.shipping_quantity * od.unit_weight                         AS shipping_weight,
    od.order_number_id
FROM shipping_detail_tb sd
JOIN shipping_tb      st ON st.shippingid       = sd.shippingid_id
LEFT JOIN driver_tb   d  ON d.driver_name       = st.driver_name_id
LEFT JOIN project_tb  p  ON p.shipping_address  = st.shipping_address
JOIN order_detail_tb  od ON od.detailid         = sd.detailid_id
JOIN order_tb         ot ON ot.order_number     = od.order_number_id
WHERE sd.shippingid_id = %s
GROUP BY
    sd.detailid_id, sd.shippingid_id, st.delivery_date, st.shipping_address,
    ot.project_name, od.commodity_details, od.commodity_size, od.commodity_units,
    od.control_no, od.unit_weight, sd.shipping_quantity, od.order_number_id;
"""


def _rows_as_dict(cur):
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def _build_excel_stream(
        rows: List[Dict],                   # 旧式写法
        shippingid: str
) -> Tuple[BytesIO, str]:
    """
    参数
    ----
    rows       : 由 SQL 查询得到的 list[dict]，至少包含下列键
                 delivery_date / shipping_address / customer_name_id /
                 project_contact_name / project_contact_phone /
                 driver_name / licence_plate / shippingid_id /
                 commodity_details / commodity_size / commodity_units /
                 shipping_quantity / control_no / unit_weight /
                 shipping_weight / order_number_id / project_name
    shippingid : 送货单号（用于拼文件名）

    返回
    ----
    (BytesIO, nice_filename)
    """
    if not rows:
        raise ValueError("rows 不能为空")

    head = rows[0]

    # ── 1. 创建工作簿 / 工作表 ───────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"

    # ── 2. 顶部抬头信息 ───────────────────────────────────
    ws["A1"], ws["B1"] = "客户名称:", head["customer_name_id"]
    ws["A2"], ws["B2"] = "收货人:",   head["project_contact_name"]
    ws["A3"], ws["B3"] = "联系电话:", head["project_contact_phone"]
    ws["A4"], ws["B4"] = "司机:",     head["driver_name"]

    ws["F1"], ws["G1"] = "收货地址:", head["shipping_address"]
    ws["F2"], ws["G2"] = "送货单号:", head["shippingid_id"]
    ws["F3"], ws["G3"] = "送货日期:", head["delivery_date"]
    ws["F4"], ws["G4"] = "车牌:",     head["licence_plate"]

    # 合并 & 对齐
    ws.merge_cells("G2:H2")
    ws.merge_cells("G3:H3")
    for cell in ("G2", "G3"):
        ws[cell].alignment = Alignment(horizontal="left", vertical="bottom")

    # ── 3. 列宽 / 行高 ────────────────────────────────────
    for col, width in zip(
        "ABCDEFGHIJ", (8, 22, 20, 8, 8, 15, 8, 8, 20, 25)
    ):
        ws.column_dimensions[col].width = width

    # 抬头行字体加粗
    # ── 4. 明细表头 ──────────────────────────────────────
    ws.append([])                       # 空一行
    columns = [
        "序号", "品名", "规格", "单位", "数量",
        "编号", "件重", "总重", "订单号", "项目名称"
    ]
    ws.append(columns)
    # 统一单元格样式
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left  = Side(style="thin"),
        right = Side(style="thin"),
        top   = Side(style="thin"),
        bottom= Side(style="thin"),
    )
    # —— 在这里给 A6~J6 设样式 —— ▼
    for cell in ws["6"]:  # openpyxl 的快捷写法，等同于 ws['A6':'J6'][0]
        cell.alignment = align_center  # 已经在前面定义
        cell.border = thin_border  # 同上
    # —— 以上新增代码 —— ▲
    # ── 5. 写入明细 ──────────────────────────────────────
    total_qty = 0
    start_row_num = ws.max_row + 1      # 明细起始行号
    for idx, r in enumerate(rows, 1):
        ws.append([
            idx,
            r["commodity_details"],
            r["commodity_size"],
            r["commodity_units"],
            r["shipping_quantity"],
            r["control_no"],
            r["unit_weight"],
            r["shipping_weight"],
            r["order_number_id"],
            r["project_name"],
        ])
        total_qty += r["shipping_quantity"]

        # 当前插入行号
        current_row = start_row_num + idx-1
        for col in range(1, 11):
            cell = ws.cell(row=current_row, column=col)
            cell.alignment = align_center
            cell.border = thin_border

    # ── 6. 合计行 ────────────────────────────────────────
    end_row = ws.max_row + 1
    ws[f"D{end_row}"] = "合计"
    ws[f"E{end_row}"] = total_qty
    ws[f"D{end_row}"].alignment = align_center
    ws[f"E{end_row}"].alignment = align_center

    # 调整行高
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    # ── 7. 保存到内存并返回 ───────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_str = head["delivery_date"].strftime("%Y-%m-%d")
    nice_name = f"{date_str} {head['shipping_address']} {shippingid}.xlsx"

    return buffer, nice_name
class GroupConcat(Aggregate):
    function = 'GROUP_CONCAT'
    output_field = CharField()
    allow_distinct = True           # 关键：告诉 Django 接受 distinct 参数
    template = (
        "%(function)s(%(distinct)s%(expressions)s "
        "SEPARATOR '%(delimiter)s')"
    )

    def __init__(self, expression, *, delimiter=',', distinct=False, **extra):
        super().__init__(
            expression,
            delimiter=delimiter,
            distinct=distinct,      # 直接透传
            **extra
        )
class ShippingTableView(object):
    def query(self, request):
        page_num = request.data.get('pageNum')
        page_size = request.data.get('pageSize')
        search_form = request.data.get('searchForm')
        search_data = search_form.get('search_data')
        role = UserTableView().getRole(search_form.get('username'))
        delivery_date = search_form.get('delivery_date')
        if role == 1:
            del search_form['username']

        if search_data is None:
            search_data=''
        order_number_subquery = (
            OrderDetailTable.objects
            .filter(
                detailid=OuterRef('shippingdetailtable__detailid'),  # 关联字段
                order_number__order_number__icontains=search_data  # 模糊匹配
            )
            .values_list('order_number__order_number', flat=True)  # 只取字段
            .distinct()  # ← 去重
            [:1]  # ← 只保留 1 行给 Subquery
        )
        # query_set = ShippingTable.objects.filter(Q(**search_form)).all().values()
        if delivery_date is None:
            # 查询 ShippingTable，并聚合出匹配的 order_number
            query_set = ShippingTable.objects.filter(
                    Q(shippingdetailtable__detailid__order_number__order_number__icontains=search_data) |
                    Q(customer_name_id__customer_name__icontains=search_data) |
                    Q(shipping_address__icontains=search_data) |
                    Q(shippingid__icontains=search_data) |
                    Q(delivery_date__icontains=search_data)
            ).annotate(
                order_number=GroupConcat(
                    'shippingdetailtable__detailid__order_number__order_number',
                    distinct=True,  # ★ 关键：聚合时再去重
                    delimiter=','
                )
            ).order_by('-delivery_date').values()

        else:
            query_set = ShippingTable.objects.filter(
                (
                    Q(shippingdetailtable__detailid__order_number__order_number__icontains=search_data) |
                    Q(customer_name_id__customer_name__icontains=search_data) |
                    Q(shipping_address__icontains=search_data) |
                    Q(shippingid__icontains=search_data) |
                    Q(delivery_date__icontains=search_data)
                ) &
                Q(delivery_date__range=(delivery_date[0], delivery_date[1]))
            ).annotate(
                # 聚合 order_number
                order_number=GroupConcat(
                    'shippingdetailtable__detailid__order_number__order_number',
                    distinct=True,  # ★ 关键：聚合时再去重
                    delimiter=','
                )
            ).order_by('-delivery_date').values()
        paginator = Paginator(query_set, page_size)
        page_data = paginator.page(page_num)
        resp = {'code': 0, 'msg': 'success', 'data': list(page_data), 'total': len(query_set)}
        return JsonResponse(resp)

    def addRecord(self, request):
        foreign_key_fields = {}
        for field in ShippingTable._meta.get_fields():
            if field.many_to_one:
                foreign_key_fields[field.name + '_id'] = field.related_model
        addForm = {}
        modalForm = request.data.get('modalForm')
        for key, value in modalForm.items():
            if key in foreign_key_fields:
                mmap = {key[:-3]: value}
                obj = foreign_key_fields[key].objects.get(**mmap)
                addForm[key[:-3]] = obj
            else:
                addForm[key] = value
        try:
            ShippingTable.objects.create(**addForm)
        except Exception as e:
            logger.error(str(e))
            resp = {'code': API_SYS_ERR, 'msg': str(e), 'data': ''}
            return JsonResponse(resp)
        resp = {'code': API_OK, 'msg': 'succ', 'data': ''}
        return JsonResponse(resp)

    def deleteRecord(self, request):
        idmap = request.data.get('idmap')
        try:
            datas = ShippingDetailTable.objects.filter(**idmap).all()
            for item_1 in datas:
                detailid = item_1.detailid.detailid
                quantity = item_1.shipping_quantity
                OrderDetailTable.objects.filter(detailid=detailid).update(
                    commodity_quantity=F("commodity_quantity") + quantity)
            ShippingTable.objects.filter(**idmap).delete()
        except Exception as e:
            logger.error(str(e))
            resp = {'code': API_SYS_ERR, 'msg': str(e), 'data': ''}
            return JsonResponse(resp)
        resp = {'code': API_OK, 'msg': 'succ', 'data': ''}
        return JsonResponse(resp)

    def editRecord(self, request):
        foreign_key_fields = {}
        for field in ShippingTable._meta.get_fields():
            if field.many_to_one:
                foreign_key_fields[field.name + '_id'] = field.related_model
        editForm = {}
        modalEditForm = request.data.get('modalEditForm')
        for key, value in modalEditForm.items():
            if key in foreign_key_fields:
                mmap = {key[:-3]: value}
                obj = foreign_key_fields[key].objects.get(**mmap)
                editForm[key[:-3]] = obj
            else:
                editForm[key] = value
        idmap = request.data.get('idmap')
        try:
            ShippingTable.objects.filter(**idmap).update(**editForm)
        except Exception as e:
            logger.error(str(e))
            resp = {'code': API_SYS_ERR, 'msg': str(e), 'data': ''}
            return JsonResponse(resp)
        resp = {'code': API_OK, 'msg': 'succ', 'data': ''}
        return JsonResponse(resp)

    def getOptions(self, request):
        data = {}
        options = []
        for item in ProjectTable.objects.all().values('shipping_address'):
            options.append({'label': item['shipping_address'], 'value': item['shipping_address']})
        data['shipping_address_options'] = options
        options = []
        for item in DriverTable.objects.all().values('driver_name'):
            options.append({'label': item['driver_name'], 'value': item['driver_name']})
        data['driver_name_options'] = options
        options = []
        for item in UserTable.objects.all().values('username'):
            options.append({'label': item['username'], 'value': item['username']})
        data['username_options'] = options
        options = []
        for item in CustomerTable.objects.all().values('customer_name'):
            options.append({'label': item['customer_name'], 'value': item['customer_name']})
        data['customer_name_options'] = options
        resp = {'code': 0, 'msg': 'success', 'data': data}
        return JsonResponse(resp)



    def __genereate_pdf(self, data,file_type,shippingid):
        if file_type == "1":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header.html')
        elif file_type == "2":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template_2.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header.html')
        elif file_type == "3":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template_3.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header.html')
        elif file_type == "4":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header_shengyuan.html')
        elif file_type == "5":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template_2.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header_shengyuan.html')
        elif file_type == "6":
            pdf_template_path = os.path.join(settings.BASE_DIR, 'templates', 'shipping_template_3.html')
            header_template_path = os.path.join(settings.BASE_DIR, 'templates', 'header_shengyuan.html')

        footer_template_path = os.path.join(settings.BASE_DIR, 'templates', 'footer.html')
        with open(pdf_template_path, 'r', encoding='utf-8') as file:
            template_content = file.read()
        template = Template(template_content)
        table_data = []
        for item in data:
            table_data.append(item)

        info = data[0]
        info['totalpages'] = math.ceil(len(table_data) / 20) + 1
        total_quantity = 0
        total_weight = 0
        for item in table_data:
            total_quantity += int(item['shipping_quantity'])
            total_weight += item['shipping_weight'] if item['shipping_weight'] is not None else 0

        info['total_quantity'] = total_quantity
        info['total_weight'] = total_weight
        picilist = _get_pici_list(shippingid)
        rendered_html_content = template.render(rows=table_data, info=info, picilist=picilist,)
        with open(header_template_path, 'r', encoding='utf-8') as file:
            template_content2 = file.read()
        with open(footer_template_path, 'r', encoding='utf-8') as file:
            template_content3 = file.read()
        template2 = Template(template_content2)
        template3 = Template(template_content3)
        header_html_path = template2.render(info=info)
        footer_html_path = template3.render(info=info)
        rendered_html_filename = os.path.join(settings.PDF_DIR, f'shipping_{shippingid}.html')
        with open(rendered_html_filename, 'w', encoding='utf-8') as file:
            file.write(rendered_html_content)
        header_html_path_filename = os.path.join(settings.PDF_DIR, f'shipping_{shippingid}2.html')
        with open(header_html_path_filename, 'w', encoding='utf-8') as file:
            file.write(header_html_path)
        footer_html_pathl_filename = os.path.join(settings.PDF_DIR, f'shipping_{shippingid}3.html')
        with open(footer_html_pathl_filename, 'w', encoding='utf-8') as file:
            file.write(footer_html_path)
        options = {
            'page-size': 'A4',
            'margin-top': '70mm',
            'margin-right': '10mm',
            'margin-bottom': '40mm',
            'margin-left': '10mm',
            'encoding': 'UTF-8',
            'no-outline': None,
            "enable-local-file-access": True,
            'header-html': header_html_path_filename,
            'footer-html':footer_html_pathl_filename,
            'header-right': '第'+'[page]'+'页 ''共'+'[topage]'+'页 ',
            'header-font-name':'宋体',
            'header-font-size':'8'
        }
        config = pdfkit.configuration(wkhtmltopdf='.\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
        row = data[0]
        # 1. 把日期转成想要的格式，例如 2025-05-17
        date_str = row['delivery_date'].strftime('%Y-%m-%d')
        # 2. 其余字段也要保证是字符串，None 时给个占位
        addr = str(row.get('shipping_address', ''))
        sid = str(row.get('shippingid_id', ''))
        # 3. 组合文件名
        pdf_filename = f"{date_str} {addr} {sid}.pdf"
        pdf_file_path = os.path.join(settings.PDF_DIR, pdf_filename)
        ret = pdfkit.from_file(rendered_html_filename, pdf_file_path, configuration=config, options=options)
        if ret:
            logger.info(f'Generate pdf file {pdf_filename} success')
        else:
            logger.error(f'Generate pdf file {pdf_filename} failed, msg: {ret}')
        print(pdf_filename)
        return pdf_filename

    def downloadRecord(self, request):
        shippingid = request.data.get('shippingid')
        print(shippingid)
        file_type = request.data.get('file_type')
        sql = f"""
                SELECT 
                    sd.detailid_id AS id,                      -- 唯一标识
                    sd.shippingid_id AS shippingid_id,         -- 送货单号
                    st.delivery_date,                          -- 送货日期
                    st.shipping_address,                       -- 收货地址
                    st.driver_name_id,                         -- 司机ID
                    GROUP_CONCAT(DISTINCT d.driver_name) AS driver_name,        -- 司机名称
                    GROUP_CONCAT(DISTINCT d.licence_plate) AS licence_plate,    -- 车牌号
                    GROUP_CONCAT(DISTINCT p.project_contact_name) AS project_contact_name,  -- 收货人
                    GROUP_CONCAT(DISTINCT p.project_contact_phone) AS project_contact_phone, -- 联系电话
                    ot.project_name,                           -- 项目名称
                    od.customer_name_id,                      -- 客户名称
                    od.commodity_details,                      -- 品名
                    od.commodity_size,                         -- 规格
                    od.commodity_units,                        -- 单位
                    od.control_no,                             -- 编号
                    od.unit_weight,                            -- 件重
                    sd.shipping_quantity,                      -- 数量
                    (sd.shipping_quantity * od.unit_weight) AS shipping_weight, -- 总重
                    od.order_number_id                         -- 订单号
                FROM 
                    shipping_detail_tb sd
                INNER JOIN 
                    shipping_tb st ON sd.shippingid_id = st.shippingid
                LEFT JOIN 
                    driver_tb d ON st.driver_name_id = d.driver_name
                LEFT JOIN 
                    project_tb p ON st.shipping_address = p.shipping_address
                INNER JOIN 
                    order_detail_tb od ON sd.detailid_id = od.detailid
                INNER JOIN 
                    order_tb ot ON od.order_number_id = ot.order_number
                WHERE 
                    sd.shippingid_id = {shippingid}            -- 替换为动态参数
                GROUP BY 
                    sd.detailid_id, sd.shippingid_id, st.delivery_date, st.shipping_address, st.driver_name_id, 
                    ot.project_name, od.commodity_details, od.commodity_size, od.commodity_units, 
                    od.control_no, od.unit_weight, sd.shipping_quantity, od.order_number_id;
        """
        code, msg, data = DBUtil.query(sql, {})
        self.__genereate_pdf(data,file_type,shippingid)
        return JsonResponse({'code': code, 'msg': msg, 'data': []})
    def downloadExcel(self, request):
        shippingid = request.data.get('shippingid')
        print(shippingid)
        sql = f"""
        SELECT 
            sd.detailid_id AS id,                      -- 唯一标识
            sd.shippingid_id AS shippingid_id,         -- 送货单号
            st.delivery_date,                          -- 送货日期
            st.shipping_address,                       -- 收货地址
            st.driver_name_id,                         -- 司机ID
            GROUP_CONCAT(DISTINCT d.driver_name) AS driver_name,        -- 司机名称
            GROUP_CONCAT(DISTINCT d.licence_plate) AS licence_plate,    -- 车牌号
            GROUP_CONCAT(DISTINCT p.project_contact_name) AS project_contact_name,  -- 收货人
            GROUP_CONCAT(DISTINCT p.project_contact_phone) AS project_contact_phone, -- 联系电话
            ot.project_name,                           -- 项目名称
            od.customer_name_id,                      -- 客户名称
            od.commodity_details,                      -- 品名
            od.commodity_size,                         -- 规格
            od.commodity_units,                        -- 单位
            od.control_no,                             -- 编号
            od.unit_weight,                            -- 件重
            sd.shipping_quantity,                      -- 数量
            (sd.shipping_quantity * od.unit_weight) AS shipping_weight, -- 总重
            od.order_number_id                         -- 订单号
        FROM 
            shipping_detail_tb sd
        INNER JOIN 
            shipping_tb st ON sd.shippingid_id = st.shippingid
        LEFT JOIN 
            driver_tb d ON st.driver_name_id = d.driver_name
        LEFT JOIN 
            project_tb p ON st.shipping_address = p.shipping_address
        INNER JOIN 
            order_detail_tb od ON sd.detailid_id = od.detailid
        INNER JOIN 
            order_tb ot ON od.order_number_id = ot.order_number
        WHERE 
            sd.shippingid_id = {shippingid}            -- 替换为动态参数
        GROUP BY 
            sd.detailid_id, sd.shippingid_id, st.delivery_date, st.shipping_address, st.driver_name_id, 
            ot.project_name, od.commodity_details, od.commodity_size, od.commodity_units, 
            od.control_no, od.unit_weight, sd.shipping_quantity, od.order_number_id;
        """
        code, msg, data = DBUtil.query(sql, {})
        self.__genereate_excel(data,shippingid)
        return JsonResponse({'code': code, 'msg': msg, 'data':[]})
    def __genereate_excel(self,data,shippingid):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet['A1']='客户名称:'
        sheet['B1']=data[0]['customer_name_id']
        sheet['A2']='收货人:'
        sheet['B2']=data[0]['project_contact_name']
        sheet['A3']='联络电话:'
        sheet['B3']=data[0]['project_contact_phone']
        sheet['A4']='送货司机:'
        sheet['B4']=data[0]['driver_name']

        sheet['F1']='收货地址:'
        sheet['G1']=data[0]['shipping_address']
        sheet['F2']='送货单号:'
        sheet['G2']=data[0]['shippingid_id']
        sheet['F3']='送货日期:'
        sheet['G3']=data[0]['delivery_date']
        sheet['F4']='车   牌:'
        sheet['G4']=data[0]['licence_plate']

        sheet['A5'] = ''

        columns = [
            '序号','品名', '规格', '单位', '数量', '编号',
            '件重', '总重', '订单号码', '项目名称'
        ]
        sheet.append(columns)
        index = 1
        total = 0
        for item in data:
            sheet.append([
                index,
                item["commodity_details"],
                item["commodity_size"],
                item["commodity_units"],
                item["shipping_quantity"],
                item["control_no"],
                item["unit_weight"],
                item["shipping_weight"],
                item["order_number_id"],
                item["project_name"]
            ])
            total = total + item["shipping_quantity"]
            index = index+1

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['D'].width = 8
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['G'].width = 8
        sheet.column_dimensions['H'].width = 8
        sheet.column_dimensions['F'].width = 15

        sheet[f'D{index+6}']='合计'
        sheet[f'E{index + 6}'] = total

        sheet.merge_cells('G2:H2')
        sheet.merge_cells('G3:H3')
        sheet['G2'].alignment = Alignment(
            horizontal='left',
            vertical='bottom'
        )
        sheet['G3'].alignment = Alignment(
            horizontal='left',
            vertical='bottom'
        )

        alignment = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0,
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row_num in range(0, index+7):
            sheet.row_dimensions[row_num].height = 20  # 设置行高为 30

        for row in sheet.iter_rows(min_row=6, max_row=index+5, min_col=1, max_col=10):
            for cell in row:
                cell.alignment = alignment
                cell.border = thin_border

        sheet[f'D{index + 6}'].alignment = alignment
        sheet[f'E{index + 6}'].alignment = alignment
        row = data[0]
        # 1. 把日期转成想要的格式，例如 2025-05-17
        date_str = row['delivery_date'].strftime('%Y-%m-%d')
        # 2. 其余字段也要保证是字符串，None 时给个占位
        addr = str(row.get('shipping_address', ''))
        sid = str(row.get('shippingid_id', ''))
        # 3. 组合文件名
        excel_file_name = f"{date_str} {addr} {sid}.xlsx"
        file_path =  os.path.join(settings.PDF_DIR, excel_file_name)
        workbook.save(file_path)
        # 确保文件存在
        # if not os.path.exists(file_path):
        #     return HttpResponse(status=404)
        # # 返回文件响应
        # responseFile = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f'shipping_{shippingid}.xlsx')
        # responseFile['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # # 设置内容类型为 Excel 文件
        print(excel_file_name)
        return excel_file_name




    def getShippingInfoByShippingId(self, request):
        """ 根据id获取发货信息 """
        shippingid = request.data.get('shippingid')
        sql = f"""
        select s.delivery_date as delivery_date,
           s.customer_name_id as customer_name,
           s.driver_name_id as driver_name,
           s.username_id as username,
           p.project_contact_name as project_contact_name,
           p.project_contact_phone as project_contact_phone,
           p.shipping_address as shipping_address,
           d.licence_plate as licence_plate
        from shipping_tb s, driver_tb d, project_tb p
        where s.driver_name_id = d.driver_name and s.shipping_address = p.shipping_address
        and shippingid = {shippingid}
        """
        db_util = DBUtil()
        code, msg, data = db_util.query(sql)
        print(data)
        return JsonResponse({'code': API_OK, 'msg': 'succ', 'data': data})
    def batchDownloadPdf(self, request):
        """
        前端传 shippingids = [10001, 10002, ...]，可选 file_type（默认为 '1'）
        返回 data = {'zip_name':'shipping_pdf_20250517_abc.zip'}
        """
        shippingids = request.data.get('shippingids', [])
        file_type   = request.data.get('file_type', '1')

        if not shippingids:
            return JsonResponse({'code': API_PARAM_ERR, 'msg': 'shippingids 不能为空', 'data': []})

        pdf_files = []           # 绝对路径
        for sid in shippingids:
            # ① 复用单条生成逻辑
            data = self.__get_detail_data(sid)      # ↓ 工具函数，见下
            pdf_name = self.__genereate_pdf(data, file_type, sid)   # 已返回文件名
            pdf_files.append(os.path.join(settings.PDF_DIR, pdf_name))

        zip_name = self.__make_zip(pdf_files, prefix='shipping_pdf')
        return JsonResponse({'code': API_OK, 'msg': 'succ', 'data': {'zip_name': zip_name}})

    # --------------------------------------------------
    # 批量下载 Excel
    # --------------------------------------------------
    def batchDownloadExcel(self, request):
        """
        与上面同理，生成多个 xlsx → 打包 zip
        """
        shippingids = request.data.get('shippingids', [])
        if not shippingids:
            return JsonResponse({'code': API_PARAM_ERR, 'msg': 'shippingids 不能为空', 'data': []})

        excel_files = []
        for sid in shippingids:
            data = self.__get_detail_data(sid)
            excel_name = self.__genereate_excel(data, sid)          # 已返回文件名
            excel_files.append(os.path.join(settings.PDF_DIR, excel_name))

        zip_name = self.__make_zip(excel_files, prefix='shipping_excel')
        return JsonResponse({'code': API_OK, 'msg': 'succ', 'data': {'zip_name': zip_name}})

    # --------------------------------------------------
    # 共用小工具：拿明细数据
    # --------------------------------------------------
    def __get_detail_data(self, shippingid: int):
        """
        把 downloadRecord / downloadExcel 中的大 SQL 提出来复用，
        仅把 shippingid 格式化进去，返回 data 列表
        """
        sql = f"""
                SELECT 
                    sd.detailid_id AS id,                      -- 唯一标识
                    sd.shippingid_id AS shippingid_id,         -- 送货单号
                    st.delivery_date,                          -- 送货日期
                    st.shipping_address,                       -- 收货地址
                    st.driver_name_id,                         -- 司机ID
                    GROUP_CONCAT(DISTINCT d.driver_name) AS driver_name,        -- 司机名称
                    GROUP_CONCAT(DISTINCT d.licence_plate) AS licence_plate,    -- 车牌号
                    GROUP_CONCAT(DISTINCT p.project_contact_name) AS project_contact_name,  -- 收货人
                    GROUP_CONCAT(DISTINCT p.project_contact_phone) AS project_contact_phone, -- 联系电话
                    ot.project_name,                           -- 项目名称
                    od.customer_name_id,                      -- 客户名称
                    od.commodity_details,                      -- 品名
                    od.commodity_size,                         -- 规格
                    od.commodity_units,                        -- 单位
                    od.control_no,                             -- 编号
                    od.unit_weight,                            -- 件重
                    sd.shipping_quantity,                      -- 数量
                    (sd.shipping_quantity * od.unit_weight) AS shipping_weight, -- 总重
                    od.order_number_id                         -- 订单号
                FROM 
                    shipping_detail_tb sd
                INNER JOIN 
                    shipping_tb st ON sd.shippingid_id = st.shippingid
                LEFT JOIN 
                    driver_tb d ON st.driver_name_id = d.driver_name
                LEFT JOIN 
                    project_tb p ON st.shipping_address = p.shipping_address
                INNER JOIN 
                    order_detail_tb od ON sd.detailid_id = od.detailid
                INNER JOIN 
                    order_tb ot ON od.order_number_id = ot.order_number
                WHERE 
                    sd.shippingid_id = {shippingid}            -- 替换为动态参数
                GROUP BY 
                    sd.detailid_id, sd.shippingid_id, st.delivery_date, st.shipping_address, st.driver_name_id, 
                    ot.project_name, od.commodity_details, od.commodity_size, od.commodity_units, 
                    od.control_no, od.unit_weight, sd.shipping_quantity, od.order_number_id;
        """
        code, msg, data = DBUtil.query(sql, {})
        if code != API_OK:
            raise RuntimeError(msg)
        return data

    # --------------------------------------------------
    # 共用小工具：把若干文件打包 ZIP
    # --------------------------------------------------
    def __make_zip(self, file_paths, *, prefix='shipping'):
        """
        file_paths: 绝对路径列表
        返回 zip 文件名（仅文件名，前端下载时再拼 settings.PDF_DIR）
        """
        zip_name = f'{prefix}_{datetime.datetime.now():%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:6]}.zip'
        zip_path = os.path.join(settings.PDF_DIR, zip_name)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fp in file_paths:
                if os.path.exists(fp):
                    zf.write(fp, arcname=os.path.basename(fp))

        return zip_name

    def make_response(shippingid: str) -> Union[StreamingHttpResponse, JsonResponse]:
        with connection.cursor() as cur:
            cur.execute(RAW_SQL, (shippingid,))  # 一元元组，防注入
            rows = _rows_as_dict(cur)

        if not rows:
            return JsonResponse({"code": 1, "msg": "no data"})

        buf, name = _build_excel_stream(rows, shippingid)  # name 例：2025-05-22 澳門 54099320.xlsx

        resp = StreamingHttpResponse(
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ── 关键修改开始 ─────────────────────────────────────────
        ascii_fallback = f"shipping_{shippingid}.xlsx"  # 万一浏览器不识别 filename*
        quoted = quote(name)  # UTF-8 → %E6%B8%AF…
        print(quote(name[:-5]))
        resp["Content-Disposition"] = (
            f'attachment; filename="{ascii_fallback}"; '
            f"filename*=UTF-8''{quoted}"
        )
        # ── 关键修改结束 ─────────────────────────────────────────

        resp["Access-Control-Expose-Headers"] = "*"
        resp["my_file_type"] = "xlsx"
        resp["my_file_name_no_suffix"] = quote(name[:-5])   # 去掉 .xlsx
        return resp

    def downloadRecord(shippingid, file_type):

        # ——— 查询（用参数化，占位符 %s）—————————————
        RAW_SQL = """
        SELECT
          sd.detailid_id           AS id,
          sd.shippingid_id         AS shippingid_id,
          st.delivery_date,
          st.shipping_address,
          GROUP_CONCAT(DISTINCT d.driver_name)        AS driver_name,
          GROUP_CONCAT(DISTINCT d.licence_plate)      AS licence_plate,
          GROUP_CONCAT(DISTINCT p.project_contact_name)  AS project_contact_name,
          GROUP_CONCAT(DISTINCT p.project_contact_phone) AS project_contact_phone,
          ot.project_name,
          od.customer_name_id,
          od.commodity_details,
          od.commodity_size,
          od.commodity_units,
          od.control_no,
          od.unit_weight,
          sd.shipping_quantity,
          sd.shipping_quantity * od.unit_weight AS shipping_weight,
          od.order_number_id
        FROM shipping_detail_tb sd
        JOIN shipping_tb      st ON st.shippingid       = sd.shippingid_id
        LEFT JOIN driver_tb   d  ON d.driver_name       = st.driver_name_id
        LEFT JOIN project_tb  p  ON p.shipping_address  = st.shipping_address
        JOIN order_detail_tb  od ON od.detailid         = sd.detailid_id
        JOIN order_tb         ot ON ot.order_number     = od.order_number_id
        WHERE sd.shippingid_id = %s
        GROUP BY
          sd.detailid_id, sd.shippingid_id, st.delivery_date, st.shipping_address,
          ot.project_name, od.commodity_details, od.commodity_size, od.commodity_units,
          od.control_no, od.unit_weight, sd.shipping_quantity, od.order_number_id
        """
        from django.db import connection
        with connection.cursor() as cur:
            cur.execute(RAW_SQL, (shippingid,))
            cols = [c[0] for c in cur.description]
            data = [dict(zip(cols, row)) for row in cur.fetchall()]

        if not data:
            return JsonResponse({'code': 1, 'msg': 'no data'})

        # ——— 生成 PDF 到 BytesIO ——————————————
        pdf_stream, nice_name = _build_pdf_stream(data, file_type, shippingid)

        resp = StreamingHttpResponse(
            pdf_stream,
            content_type='application/pdf'
        )
        fallback = f"shipping_{shippingid}.pdf"
        quoted = quote(nice_name)  # RFC 5987
        resp['Content-Disposition'] = (
            f'attachment; filename="{fallback}"; '
            f"filename*=UTF-8''{quoted}"
        )
        resp['Access-Control-Expose-Headers'] = '*'
        resp['my_file_type'] = 'pdf'
        resp['my_file_name_no_suffix'] = quote(nice_name[:-4])
        return resp

    # ---------- 把原 __genereate_pdf 改成只返回 BytesIO ----------
def _build_pdf_stream(
           data: List[Dict],  # ⬅️ 旧写法 list[dict] → 新写法 List[Dict]
           file_type: str,
           shippingid: str
   ) -> Tuple[BytesIO, str]:  # ⬅️ 返回 (BytesIO, 文件名)
       """返回 (BytesIO(), nice_filename)"""
       # 1. 选模板
       base = os.path.join(settings.BASE_DIR, 'templates')
       header_tmpl = 'header.html'
       if file_type in ('4', '5', '6'):
           header_tmpl = 'header_shengyuan.html'
       body_map = {'1': 'shipping_template.html',
                   '2': 'shipping_template_2.html',
                   '3': 'shipping_template_3.html',
                   '4': 'shipping_template.html',
                   '5': 'shipping_template_2.html',
                   '6': 'shipping_template_3.html'}
       body_tmpl = body_map.get(file_type, 'shipping_template.html')

       footer_tmpl = 'footer.html'

      # 2. 渲染 Jinja2
       def _render(path, ctx):
           with open(os.path.join(base, path), 'r', encoding='utf-8') as f:
               return Template(f.read()).render(**ctx)

       table_data = data
       info = data[0].copy()
       info['totalpages'] = math.ceil(len(table_data) / 20) + 1
       info['total_quantity'] = sum(int(r['shipping_quantity']) for r in table_data)
       info['total_weight'] = sum(r['shipping_weight'] or 0 for r in table_data)
       picilist = _get_pici_list(shippingid)
       html_body = _render(body_tmpl, {'rows': table_data, 'info': info, 'picilist': picilist})
       html_header = _render(header_tmpl, {'info': info})
       html_footer = _render(footer_tmpl, {'info': info})

       # 3. 直接把 html 字符串传 pdfkit → BytesIO
       options = {
           'page-size': 'A4',
           'margin-top': '70mm',
           'margin-right': '10mm',
           'margin-bottom': '40mm',
           'margin-left': '10mm',
           'encoding': 'UTF-8',
           'no-outline': None,
           'enable-local-file-access': None,
           'header-html': '-',  # 占位符，稍后通过 'input' 传
           'footer-html': '-',
           'header-right': '第[page]页 共[topage]页 ',
           'header-font-name': '宋体',
           'header-font-size': '8'
       }
       config = pdfkit.configuration(wkhtmltopdf=r'.\wkhtmltopdf\bin\wkhtmltopdf.exe')

       # html_to_pdf 需要把 header/footer 临时文件写磁盘
       import tempfile, pathlib
       with tempfile.TemporaryDirectory() as tmpdir:
           header_file = pathlib.Path(tmpdir, 'h.html');
           header_file.write_text(html_header, encoding='utf-8')
           footer_file = pathlib.Path(tmpdir, 'f.html');
           footer_file.write_text(html_footer, encoding='utf-8')
           options['header-html'] = str(header_file)
           options['footer-html'] = str(footer_file)

           pdf_bytes = pdfkit.from_string(
               html_body,
               False,  # ← 关键：False 表示“返回 bytes，而不是写文件”
               configuration=config,
               options=options
           )

           # ② 包装成 BytesIO，供 StreamingHttpResponse 使用
           pdf_io = BytesIO(pdf_bytes)
           pdf_io.seek(0)
           # 4. 生成下载文件名
       date_str = info['delivery_date'].strftime('%Y-%m-%d')
       nice_name = f"{date_str} {info['shipping_address']} {shippingid}.pdf"
       return pdf_io, nice_name
def _get_pici_list(shippingid):
    sql = f"""
            select r.order_number, count(*) as num
        from (
                select o.order_number
                from shipping_tb s, shipping_detail_tb sd, order_detail_tb d, order_tb o
                where s.shippingid = sd.shippingid_id and sd.detailid_id = d.detailid and d.order_number_id = o.order_number
                group by s.shippingid, o.order_number
                    ) r
        where r.order_number in (
            select o.order_number from shipping_tb s, order_detail_tb d, order_tb o, shipping_detail_tb sd
                where s.shippingid = {shippingid} and sd.shippingid_id = {shippingid} and sd.detailid_id = d.detailid
                and d.detailid = sd.detailid_id and d.order_number_id = o.order_number
                group by o.order_number
            )
        group by r.order_number
    """
    sql_query2 = f"""
        SELECT
            r.order_number,
            CASE
                WHEN COALESCE(SUM(od.commodity_quantity), 0) = 0 THEN CONCAT('(完)')
                ELSE ''
            END AS order_number_with_suffix
        FROM (
            SELECT DISTINCT o.order_number
            FROM shipping_tb s
            JOIN shipping_detail_tb sd ON s.shippingid = sd.shippingid_id
            JOIN order_detail_tb d ON sd.detailid_id = d.detailid
            JOIN order_tb o ON d.order_number_id = o.order_number
            WHERE s.shippingid = {shippingid}
        ) r
        LEFT JOIN order_detail_tb od ON r.order_number = od.order_number_id
        GROUP BY r.order_number;
    """
    code1, msg1, data1 = DBUtil.query(sql, {})
    code2, msg2, data2 = DBUtil.query(sql_query2, {})
    combined_data = {}
    for item in data1:
        order_number = item['order_number']
        combined_data[order_number] = {
            'order_number': order_number,
            'num': item['num']
        }

    # 处理第二段数据
    for item in data2:
        order_number = item['order_number']
        if order_number in combined_data:
            combined_data[order_number]['order_number_with_suffix'] = item['order_number_with_suffix']
        else:
            combined_data[order_number] = {
                'order_number': order_number,
                'order_number_with_suffix': item['order_number_with_suffix']
            }

    # 将字典转换为列表
    newdata = [{'order_number': k, **v} for k, v in combined_data.items()]
    print(newdata)
    return newdata